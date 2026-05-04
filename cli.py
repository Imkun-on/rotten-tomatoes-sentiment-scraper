"""CLI interface for Review Scraper using Rich."""

import csv
import json
import os
import sys

from rich.align import Align
from rich.box import DOUBLE, ROUNDED
from rich.console import Console
from rich.panel import Panel
from rich.progress import (
    Progress, SpinnerColumn, BarColumn, TextColumn,
    MofNCompleteColumn, TimeElapsedColumn, TimeRemainingColumn,
    TaskProgressColumn,
)
from rich.prompt import Prompt, IntPrompt
from rich.style import Style
from rich.table import Table
from rich.text import Text
from rich.theme import Theme

from models import MovieInfo, ReviewData, SeasonInfo
from scraper_rt import RTScraper
from sentiment import analyze_reviews


# ------------------------------------------------------------------
# Theme & symbols (same aesthetic as scraper_anime)
# ------------------------------------------------------------------

CUSTOM_THEME = Theme({
    "info": "bright_cyan",
    "success": "bright_green",
    "warning": "yellow",
    "error": "bold red",
    "title": "bold bright_white",
    "phase": "bold bright_yellow",
    "dim_label": "dim bright_white",
})

console = Console(theme=CUSTOM_THEME)

SYM_OK = "\u2714"      # ✔
SYM_FAIL = "\u2716"    # ✖
SYM_ARROW = "\u2192"   # →
SYM_DOT = "\u2022"     # •


# ------------------------------------------------------------------
# First-prompt resilience
# ------------------------------------------------------------------

# Windows + PowerShell occasionally injects a phantom KeyboardInterrupt on
# the very first stdin read of a fresh Python session. This wrapper retries
# the first prompt of the session exactly once, then becomes a passthrough,
# so genuine Ctrl+C continues to work for every subsequent prompt.
_FIRST_PROMPT_DONE = False


def _ask_resilient_first(prompt_method, *args, **kwargs):
    global _FIRST_PROMPT_DONE
    if _FIRST_PROMPT_DONE:
        return prompt_method(*args, **kwargs)
    try:
        result = prompt_method(*args, **kwargs)
    except KeyboardInterrupt:
        result = prompt_method(*args, **kwargs)
    _FIRST_PROMPT_DONE = True
    return result


# ------------------------------------------------------------------
# Banner
# ------------------------------------------------------------------

def _print_banner() -> None:
    """Print startup banner with magenta->cyan gradient."""
    banner_lines = [
        r"    ____        __  __                ______                      __                     ____            _                  ",
        r"   / __ \____  / /_/ /____  ____     /_  __/___  ____ ___  ____ _/ /_____  ___  _____   / __ \___ _   __(_)__ _      _______",
        r"  / /_/ / __ \/ __/ __/ _ \/ __ \     / / / __ \/ __ `__ \/ __ `/ __/ __ \/ _ \/ ___/  / /_/ / _ \ | / / / _ \ | /| / / ___/",
        r" / _, _/ /_/ / /_/ /_/  __/ / / /    / / / /_/ / / / / / / /_/ / /_/ /_/ /  __(__  )  / _, _/  __/ |/ / /  __/ |/ |/ (__  ) ",
        r"/_/ |_|\____/\__/\__/\___/_/ /_/    /_/  \____/_/ /_/ /_/\__,_/\__/\____/\___/____/  /_/ |_|\___/|___/_/\___/|__/|__/____/  ",
        r"                                                                                                                            ",
    ]
    colors = ["bright_magenta", "magenta", "bright_blue", "blue", "bright_cyan", "cyan"]
    text = Text()
    for i, line in enumerate(banner_lines):
        text.append(line + "\n", style=Style(color=colors[i % len(colors)], bold=True))

    console.print()
    console.print(Panel(
        Align.center(text),
        border_style="bright_blue",
        box=DOUBLE,
        padding=(1, 2),
        expand=False,
    ))


# ------------------------------------------------------------------
# Display helpers
# ------------------------------------------------------------------

def show_search_results(results):
    """Display search results in a styled table."""
    table = Table(
        title=f"Search Results ({len(results)})",
        border_style="bright_cyan",
        box=ROUNDED,
        title_style="bold bright_cyan",
        row_styles=["", "dim"],
    )
    table.add_column("#", style="bold bright_white", justify="center", width=4)
    table.add_column("Title", style="info", min_width=25)
    table.add_column("Year", justify="center", style="bright_yellow", width=8)
    table.add_column("Score", justify="center", style="bold", width=10)

    for i, r in enumerate(results, 1):
        table.add_row(str(i), r.title, r.year or "", r.score or "")

    console.print(table)


def show_movie_info(info: MovieInfo):
    """Display movie info inside a rounded panel with bullet list."""
    details: list[str] = []

    details.append(f"  {SYM_DOT} [dim_label]Title:[/dim_label]  [info]{info.title}[/info]")
    if info.year:
        details.append(f"  {SYM_DOT} [dim_label]Year:[/dim_label]   [info]{info.year}[/info]")
    if info.genre:
        details.append(f"  {SYM_DOT} [dim_label]Genre:[/dim_label]  [dim]{info.genre}[/dim]")
    if info.plot:
        plot = info.plot[:1000] + "..." if len(info.plot) > 1000 else info.plot
        details.append(f"  {SYM_DOT} [dim_label]Plot:[/dim_label]   [dim]{plot}[/dim]")

    if info.director:
        details.append(f"  {SYM_DOT} [dim_label]Director:[/dim_label]     [info]{info.director}[/info]")
    if info.screenwriter:
        details.append(f"  {SYM_DOT} [dim_label]Screenwriter:[/dim_label] [dim]{info.screenwriter}[/dim]")
    if info.cast:
        details.append(f"  {SYM_DOT} [dim_label]Cast:[/dim_label]         [dim]{', '.join(info.cast[:5])}[/dim]")
    if info.runtime:
        details.append(f"  {SYM_DOT} [dim_label]Runtime:[/dim_label]      [dim]{info.runtime}[/dim]")
    if info.content_rating:
        details.append(f"  {SYM_DOT} [dim_label]Rating:[/dim_label]       [dim]{info.content_rating}[/dim]")

    if info.tomatometer:
        details.append(f"  {SYM_DOT} [bold red]Tomatometer:[/bold red]  [bright_yellow]{info.tomatometer}[/bright_yellow]  [dim]({info.tomatometer_count})[/dim]")
    if info.popcornmeter:
        details.append(f"  {SYM_DOT} [bold yellow]Popcornmeter:[/bold yellow] [bright_yellow]{info.popcornmeter}[/bright_yellow]  [dim]({info.popcornmeter_count})[/dim]")

    if info.release_date_theaters:
        details.append(f"  {SYM_DOT} [dim_label]Release (Theaters):[/dim_label]  [dim]{info.release_date_theaters}[/dim]")
    if info.release_date_streaming:
        details.append(f"  {SYM_DOT} [dim_label]Release (Streaming):[/dim_label] [dim]{info.release_date_streaming}[/dim]")
    if info.box_office:
        details.append(f"  {SYM_DOT} [dim_label]Box Office:[/dim_label]          [dim]{info.box_office}[/dim]")
    if info.budget:
        details.append(f"  {SYM_DOT} [dim_label]Budget:[/dim_label]              [dim]{info.budget}[/dim]")
    if info.country:
        details.append(f"  {SYM_DOT} [dim_label]Country:[/dim_label]             [dim]{info.country}[/dim]")
    if info.original_language:
        details.append(f"  {SYM_DOT} [dim_label]Language:[/dim_label]            [dim]{info.original_language}[/dim]")

    console.print(Panel(
        "\n".join(details),
        title=f"[title]{info.title}[/title]",
        border_style="bright_cyan",
        box=ROUNDED,
        expand=False,
        padding=(1, 3),
    ))


def show_seasons(seasons: list[SeasonInfo]):
    """Display the list of TV seasons in a styled table."""
    table = Table(
        title=f"Seasons ({len(seasons)})",
        border_style="bright_cyan",
        box=ROUNDED,
        title_style="bold bright_cyan",
        row_styles=["", "dim"],
    )
    table.add_column("#", style="bold bright_white", justify="center", width=4)
    table.add_column("Season", style="info")
    table.add_column("Episodes", justify="center", style="bright_yellow")
    table.add_column("Tomatometer", justify="center", style="bold red")
    table.add_column("Popcornmeter", justify="center", style="bold yellow")
    table.add_column("Total", justify="right", style="bright_green")

    # Pre-format each meter as "{pct:>w} ({cnt:>w})" with widths derived from
    # the widest value in the column. With every cell formatted to the same
    # width, center-justify keeps the % and the count visually aligned.
    def _fmt_pairs(pct_attr: str, num_attr: str):
        pcts, cnts = [], []
        for s in seasons:
            pct = getattr(s, pct_attr)
            num = getattr(s, num_attr)
            pcts.append(pct if pct and pct != "N/A" else "—")
            cnts.append(f"{num:,}" if num else "—")
        wp = max(len(p) for p in pcts)
        wc = max(len(c) for c in cnts)
        return [f"{p:>{wp}} ({c:>{wc}})" for p, c in zip(pcts, cnts)]

    tomato_cells = _fmt_pairs("tomatometer", "tomatometer_num")
    popcorn_cells = _fmt_pairs("popcornmeter", "popcornmeter_num")

    for i, s in enumerate(seasons, 1):
        eps = str(s.episode_count) if s.episode_count else "—"
        total = s.tomatometer_num + s.popcornmeter_num
        total_str = f"{total:,}" if total else "—"
        table.add_row(
            str(i), s.name, eps,
            tomato_cells[i - 1], popcorn_cells[i - 1],
            total_str,
        )

    console.print(table)



def parse_season_selection(spec: str, total: int) -> list[int]:
    """Parse '1', '1-3', '1,3,5', '1-2,4', 'all' into a sorted list of indices.

    Indices are 1-based positions into the season list, NOT season numbers.
    Invalid tokens are ignored; out-of-range indices are clamped.
    """
    spec = (spec or "").strip().lower()
    if not spec or spec in ("all", "a", "*"):
        return list(range(1, total + 1))

    picked: set[int] = set()
    for token in spec.split(","):
        token = token.strip()
        if not token:
            continue
        if "-" in token:
            try:
                a, b = token.split("-", 1)
                lo, hi = int(a), int(b)
                if lo > hi:
                    lo, hi = hi, lo
                for i in range(max(1, lo), min(total, hi) + 1):
                    picked.add(i)
            except ValueError:
                continue
        else:
            try:
                i = int(token)
                if 1 <= i <= total:
                    picked.add(i)
            except ValueError:
                continue
    return sorted(picked)


def show_reviews(reviews: list[ReviewData], page_size: int = 10):
    """Display reviews with pagination."""
    if not reviews:
        console.print("[warning]No reviews found.[/warning]")
        return

    total_pages = (len(reviews) + page_size - 1) // page_size
    page = 0

    while True:
        start = page * page_size
        end = min(start + page_size, len(reviews))
        page_reviews = reviews[start:end]

        console.print()
        console.rule(
            f"[phase]Reviews {start + 1}-{end} of {len(reviews)}  "
            f"(page {page + 1}/{total_pages})[/phase]",
            style="bright_blue",
        )
        console.print()

        for rev in page_reviews:
            rating_style = (
                "bright_green" if rev.rating_value >= 3.5
                else "bright_yellow" if rev.rating_value >= 2.0
                else "bright_red"
            )
            header = Text()
            header.append(f"  {SYM_DOT} ")
            header.append(f"{rev.author}", style="bold bright_white")
            header.append(f"  [{rev.rating}]", style=rating_style)
            header.append(f"  {rev.date}", style="dim")
            if rev.review_type:
                header.append(f"  ({rev.review_type})", style="dim italic")

            console.print(header)
            if rev.sentiment_label and rev.sentiment_label != "Too Short":
                sa_style = (
                    "bright_green" if rev.sentiment_score >= 4
                    else "bright_yellow" if rev.sentiment_score >= 3
                    else "bright_red"
                )
                console.print(
                    f"      [{sa_style}]Sentiment: {rev.sentiment_label} "
                    f"({rev.sentiment_score:.1f}/5)[/{sa_style}]"
                )
            if rev.text:
                text = rev.text[:300] + "..." if len(rev.text) > 300 else rev.text
                console.print(f"      [dim]{text}[/dim]")
            console.print()

        # Navigation
        options = []
        if page < total_pages - 1:
            options.append("[bold]n[/bold]ext")
        if page > 0:
            options.append("[bold]p[/bold]rev")
        options.append("[bold]q[/bold]uit")

        choice = Prompt.ask(f"  {' / '.join(options)}", default="q")
        if choice.lower() == "n" and page < total_pages - 1:
            page += 1
        elif choice.lower() == "p" and page > 0:
            page -= 1
        elif choice.lower() == "q":
            break


# ------------------------------------------------------------------
# Export
# ------------------------------------------------------------------

EXPORT_DIR = "exports"


def _ensure_export_dir():
    """Create the exports directory if it doesn't exist."""
    os.makedirs(EXPORT_DIR, exist_ok=True)


def export_reviews(reviews: list[ReviewData], movie_info: MovieInfo):
    """Export reviews to CSV, JSON, or Excel."""
    console.print()
    console.print("[title]Export format[/title]")
    console.print(f"  [bold]1[/bold]  {SYM_ARROW}  CSV")
    console.print(f"  [bold]2[/bold]  {SYM_ARROW}  JSON")
    console.print(f"  [bold]3[/bold]  {SYM_ARROW}  Excel (.xlsx)")
    console.print(f"  [bold]4[/bold]  {SYM_ARROW}  Cancel")

    choice = Prompt.ask("  Choose", choices=["1", "2", "3", "4"], default="4")
    if choice == "4":
        return

    _ensure_export_dir()

    filename = Prompt.ask(
        "  Filename (without extension)",
        default=f"{movie_info.title.replace(' ', '_')}_reviews",
    )

    COLUMNS = [
        "title", "genre", "author", "rating", "rating_value",
        "sentiment_score", "sentiment_label",
        "text", "review_type", "full_review_url", "date",
    ]

    def _review_row(r):
        return [
            r.title, r.genre, r.author, r.rating, r.rating_value,
            r.sentiment_score, r.sentiment_label,
            r.text, r.review_type, r.full_review_url, r.date,
        ]

    if choice == "1":
        path = os.path.join(EXPORT_DIR, f"{filename}.csv")
        with open(path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(COLUMNS)
            for r in reviews:
                writer.writerow(_review_row(r))
        console.print(f"  {SYM_OK} [success]Saved to {path}[/success]")

    elif choice == "2":
        path = os.path.join(EXPORT_DIR, f"{filename}.json")
        data = [dict(zip(COLUMNS, _review_row(r))) for r in reviews]
        with open(path, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
        console.print(f"  {SYM_OK} [success]Saved to {path}[/success]")

    elif choice == "3":
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            console.print(
                f"  {SYM_FAIL} [error]openpyxl non installato. "
                "Esegui: pip install openpyxl[/error]"
            )
            return

        path = os.path.join(EXPORT_DIR, f"{filename}.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.title = "Reviews"

        # Header style
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

        for col_idx, col_name in enumerate(COLUMNS, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Data rows
        for row_idx, r in enumerate(reviews, 2):
            for col_idx, value in enumerate(_review_row(r), 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Auto-fit column widths (approximate)
        for col_idx, col_name in enumerate(COLUMNS, 1):
            max_len = len(col_name)
            for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if cell.value:
                        max_len = max(max_len, min(len(str(cell.value)), 50))
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max_len + 2

        wb.save(path)
        console.print(f"  {SYM_OK} [success]Saved to {path}[/success]")


# ------------------------------------------------------------------
# Main flow
# ------------------------------------------------------------------

def main():
    _print_banner()

    while True:
        scraper = RTScraper()

        try:
            # Search mode
            console.print()
            console.print("[title]Search mode[/title]")
            console.print(f"  [bold]1[/bold]  {SYM_ARROW}  Search by query")
            console.print(f"  [bold]2[/bold]  {SYM_ARROW}  Paste a Rotten Tomatoes URL")
            mode = _ask_resilient_first(
                Prompt.ask, "  Choose", choices=["1", "2"], default="1",
            )

            if mode == "1":
                # Search by query
                console.print()
                console.rule(f"[phase]{SYM_ARROW} Search on Rotten Tomatoes[/phase]", style="bright_blue")
                query = Prompt.ask("  [bold]Query[/bold]")
                if not query.strip():
                    continue

                with console.status("[info]Searching Rotten Tomatoes...[/info]", spinner="dots"):
                    results = scraper.search(query)

                if not results:
                    console.print("[warning]No results found.[/warning]")
                    continue

                console.print()
                show_search_results(results)

                console.print()
                idx = IntPrompt.ask("  Select a result (number)", default=1)
                if idx < 1 or idx > len(results):
                    console.print("[error]Invalid selection.[/error]")
                    continue

                selected_url = results[idx - 1].url
            else:
                # Direct URL
                console.print()
                selected_url = Prompt.ask(f"  [bold]Rotten Tomatoes URL[/bold]")
                if not selected_url.strip():
                    continue

            # Movie info
            with console.status("[info]Fetching movie info...[/info]", spinner="dots"):
                movie_info = scraper.get_movie_info(selected_url)

            console.print()
            show_movie_info(movie_info)

            # TV show: pick seasons before scraping.
            selected_seasons: list[SeasonInfo] | None = None
            if "/tv/" in selected_url and "/s" not in selected_url.split("/tv/", 1)[1]:
                with console.status("[info]Fetching seasons...[/info]", spinner="dots"):
                    seasons = scraper.get_seasons(selected_url)

                if not seasons:
                    console.print("[warning]No seasons found for this TV show.[/warning]")
                    continue

                with console.status(
                    f"[info]Loading details for {len(seasons)} seasons...[/info]",
                    spinner="dots",
                ):
                    scraper.enrich_seasons(seasons)

                console.print()
                show_seasons(seasons)
                console.print(
                    f"  [dim]Selection syntax: [bold]1[/bold] | "
                    f"[bold]1-3[/bold] | [bold]1,3,5[/bold] | [bold]all[/bold][/dim]"
                )
                spec = Prompt.ask("  Choose season(s)", default="all")
                indices = parse_season_selection(spec, len(seasons))
                if not indices:
                    console.print("[error]Invalid selection.[/error]")
                    continue
                selected_seasons = [seasons[i - 1] for i in indices]

                names = ", ".join(s.name for s in selected_seasons)
                console.print(f"  {SYM_OK} [success]Selected:[/success] {names}")

            # Scrape reviews?
            console.print()
            do_scrape = Prompt.ask("  Scrape reviews?", choices=["y", "n"], default="y")
            if do_scrape == "n":
                continue

            # Scrape
            console.print()
            console.rule("[phase]\u2699 Scraping reviews[/phase]", style="bright_green")
            console.print()

            all_reviews: list[ReviewData] = []

            with Progress(
                SpinnerColumn("dots", style="bright_green"),
                TextColumn("[bold bright_green]{task.description}"),
                BarColumn(
                    bar_width=50,
                    style="bar.back",
                    complete_style="bright_green",
                    finished_style="bold green",
                ),
                TaskProgressColumn(),
                MofNCompleteColumn(),
                TextColumn("[dim]\u2502[/dim]"),
                TimeElapsedColumn(),
                TextColumn("[dim]\u2192[/dim]"),
                TimeRemainingColumn(),
                console=console,
                expand=False,
            ) as progress:
                task = progress.add_task("Scraping", total=None)

                def on_progress(current, total):
                    progress.update(task, completed=current, total=total)

                def on_batch(batch):
                    all_reviews.extend(batch)

                # Reset — on_batch already collects, so use separate list
                all_reviews.clear()
                reviews = scraper.scrape_reviews(
                    selected_url,
                    movie_info=movie_info,
                    seasons=selected_seasons,
                    on_progress=on_progress,
                )
                if not all_reviews:
                    all_reviews = reviews

            # Sentiment analysis?
            console.print()
            do_sentiment = Prompt.ask(
                "  Run sentiment analysis?", choices=["y", "n"], default="y"
            )
            if do_sentiment == "y":
                console.print()
                console.rule("[phase]🧠 Sentiment Analysis[/phase]", style="bright_magenta")
                console.print()

                with Progress(
                    SpinnerColumn("dots", style="bright_magenta"),
                    TextColumn("[bold bright_magenta]{task.description}"),
                    BarColumn(
                        bar_width=50,
                        style="bar.back",
                        complete_style="bright_magenta",
                        finished_style="bold magenta",
                    ),
                    TaskProgressColumn(),
                    MofNCompleteColumn(),
                    TextColumn("[dim]\u2502[/dim]"),
                    TimeElapsedColumn(),
                    TextColumn("[dim]\u2192[/dim]"),
                    TimeRemainingColumn(),
                    console=console,
                    expand=False,
                ) as progress:
                    sa_task = progress.add_task(
                        "Analyzing sentiment", total=len(all_reviews)
                    )

                    def on_sa_progress(current, total):
                        progress.update(sa_task, completed=current, total=total)

                    analyze_reviews(all_reviews, on_progress=on_sa_progress)

                # Sentiment summary
                labeled = [r for r in all_reviews if r.sentiment_label and r.sentiment_label != "Too Short"]
                if labeled:
                    avg = sum(r.sentiment_score for r in labeled) / len(labeled)
                    dist = {}
                    for r in labeled:
                        dist[r.sentiment_label] = dist.get(r.sentiment_label, 0) + 1

                    sa_table = Table(show_header=False, box=None, padding=(0, 2), expand=False)
                    sa_table.add_column("Label", style="dim_label")
                    sa_table.add_column("Value")
                    sa_table.add_row("Avg sentiment", f"[bright_yellow]{avg:.2f} / 5[/bright_yellow]")
                    sa_table.add_row("Analyzed", f"[success]{len(labeled)}[/success] / {len(all_reviews)}")
                    for lbl in ["Very Positive", "Positive", "Neutral", "Negative", "Very Negative"]:
                        if lbl in dist:
                            sa_table.add_row(f"  {lbl}", str(dist[lbl]))

                    console.print()
                    console.print(Panel(
                        sa_table,
                        title="🧠 Sentiment Summary",
                        border_style="bright_magenta",
                        box=DOUBLE,
                        expand=False,
                        padding=(1, 3),
                    ))

            # Summary panel
            summary = Table(show_header=False, box=None, padding=(0, 2), expand=False)
            summary.add_column("Label", style="dim_label")
            summary.add_column("Value")
            summary.add_row(f"{SYM_OK} Reviews scraped", f"[success]{len(all_reviews)}[/success]")
            summary.add_row("Movie", f"[title]{movie_info.title}[/title]")

            console.print()
            console.print(Panel(
                summary,
                title=f"{SYM_OK} Summary",
                border_style="bright_green",
                box=DOUBLE,
                expand=False,
                padding=(1, 3),
            ))

            # Show reviews
            show_reviews(all_reviews)

            # Export
            console.print()
            do_export = Prompt.ask("  Export reviews?", choices=["y", "n"], default="n")
            if do_export == "y":
                export_reviews(all_reviews, movie_info)

        except KeyboardInterrupt:
            console.print("\n[warning]Interrupted.[/warning]")
            break
        except Exception as e:
            console.print(Panel(
                f"{SYM_FAIL} [error]{e}[/error]",
                title="Error",
                border_style="red",
                box=ROUNDED,
                expand=False,
                padding=(1, 3),
            ))

        finally:
            scraper.close()

        # Continue?
        console.print()
        again = Prompt.ask("  Search again?", choices=["y", "n"], default="y")
        if again == "n":
            break

    console.print()
    console.print(Panel(
        "[title]Bye![/title]",
        border_style="bright_magenta",
        box=ROUNDED,
        expand=False,
        padding=(0, 3),
    ))


if __name__ == "__main__":
    main()
